import openpyxl

NEW_PRICES = {
    'Celery': 1.19,
    'Garlic': 3.07,
    'Lemon': 1.27
}

print('Opening workbook...')
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']


for rowNum in range(2, sheet.max_row):
    produce = sheet.cell(row=rowNum, column=1).value
#Iterate every row but the header
    if produce in NEW_PRICES:
    #Checks if the produce need to be changed
        sheet.cell(row=rowNum, column=2).value = NEW_PRICES[produce]
        sheet.cell(row=rowNum, column=4).value = round(sheet.cell(row=rowNum, column=3).value  * NEW_PRICES[produce], 2)
        
wb.save('updatedProduceSales.xlsx')